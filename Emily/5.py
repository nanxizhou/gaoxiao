from openpyxl import load_workbook
import smtplib
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart

# 设置邮箱账号
account = input('请输入邮箱账户：')
# 设置邮箱授权码
token = input('请输入邮箱授权码：')

# 实例化smtp对象，设置邮箱服务器，端口
smtp = smtplib.SMTP_SSL('smtp.qq.com',465)
# 登录qq邮箱 jlrnctylmtmzdejg
smtp.login(account,token)

# 打开工作表
wb = load_workbook('./04_月考勤表.xlsx')
sheet = wb.active

# 编写正文内容
content = '四月的考勤表已出，其中迟到时长超出 45 分钟的人员如下：\n'
for row_data in sheet.iter_rows(min_row=2, values_only=True):
    # 获取迟到时长超过45分钟的人员
    if row_data[2] > 45:
        content += '姓名：{name} 迟到总时长：{time} \n'.format(name=row_data[1], time=row_data[2])
content += '详情见附件内容'

# 设置正文，创建简单邮件对象
email_content = MIMEText(content, 'plain', 'utf-8')

# 读取工作表文件数据
with open('./04_月考勤表.xlsx', 'rb') as f:
    file_data = f.read()

# 设置内容类型为附件
attachment = MIMEText(file_data, 'base64', 'utf-8')

# 设置附件标题以及文件类型
attachment.add_header('Content-Disposition','attachment', filename='./04_月考勤表.xlsx')

# 创建复合邮件对象
msg = MIMEMultipart()

# 添加正文到复合邮件对象中
msg.attach(email_content)

# 添加附件到复合邮件对象里
msg.attach(attachment)

# 设置发送者信息
msg['From'] = '陈知枫'
# 设置接受者信息
msg['To'] = '闪光金融的各位同事们'
# 设置邮件标题
msg['Subject'] = '04_月考勤表'

# 发送邮件
smtp.sendmail(account, '2818813711@qq.com', msg.as_string())
# 关闭邮箱服务
smtp.quit()