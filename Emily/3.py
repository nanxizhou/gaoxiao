from openpyxl import load_workbook

# 打开工作簿
wb = load_workbook('./04_月考勤表.xlsx')
# 打开表格
ws = wb.active

# 使用 for 循环配合 iter_rows() 遍历表格，打印该表格中的所有人名
for row in ws.iter_rows(m):
    print(row)