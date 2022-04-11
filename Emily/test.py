from openpyxl import load_workbook
import smtplib
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart

# 设置邮箱账号
account = input('请输入邮箱: ')
# 设置邮箱授权码

# 实例化smtp对象，设置邮箱服务器，端口
smtp = smtplib.SMTP_SSL('smtp.qq.com', 465)
# 登录qq邮箱
smtp.login(account,'jlrnctylmtmzdejg')
# 打开工作表
wb = load_workbook('./04_月考勤表.xlsx', data_only=True)
sheet = wb.active

# 编写正文内容
content = '四月的考勤表已出，其中迟到时长超出 45 分钟的人员如下：\n'
for row_data in sheet.iter_rows(min_row=2, values_only=True):
    # 获取迟到时长超过45分钟的人员
    if row_data[2] > 45:
        content += '姓名：{name} 迟到总时长：{time} \n'.format(name=row_data[1], time=row_data[2])
content += '详情见附件内容'

# 设置邮件正文，实例化简单邮件对象
email_content = MIMEText(content,'plain','utf-8')
# 实例化复合邮件对象
msg = MIMEMultipart()
# 添加正文到邮件内容中
msg.attach(email_content)
# 读取工作表文件数据
with open('./04_月考勤表.xlsx', 'rb') as f:
    file_data = f.read()

# 添加附件，实例化简单邮件对象
contents = MIMEText(file_data,'base64','utf-8')
contents.add_header('Content-Disposition','attachment',filename='./04_月考勤表.xlsx')
# 将附件添加到邮件内容里
msg.attach(contents)
# 设置发送者信息
msg['From']='陈志峰'
# 设置接受者信息
msg['To'] ='王振义'
# 设置邮件标题
msg['Subject'] ='王振义的迟到记录'
# 发送邮件
smtp.sendmail(account,'2818813711@qq.com',msg.as_string())
# 关闭邮箱服务
smtp.quit()