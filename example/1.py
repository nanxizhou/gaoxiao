# 导入openpyxl模块下的的load_workbook函数
from openpyxl import load_workbook
# 文件存储路径
file_path = './material/事业01部_副本.xlsx'
# 打开工作簿
wb = load_workbook(file_path)
# 打开活动工作表
ws = wb.active