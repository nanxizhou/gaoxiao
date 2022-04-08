# 导入模块
from openpyxl import load_workbook
# 打开工作表
file_path = './material/事业01部_副本.xlsx'
wb = load_workbook(file_path)
ws = wb.active

# 打印工作表第二行所有单元格的值
for cell in ws.iter_rows(min_row=2,values_only=True):
    print( cell    )