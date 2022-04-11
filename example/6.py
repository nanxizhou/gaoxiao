# 导入模块
# 1）调整列宽，A列改为10，B列改为25，C列改为50，D列改为10，E列改为20，F列改为15。
# 2）表头单元格要用橙色纯色填充，表中单元格用淡黄色纯色填充，表尾单元格用淡桔红色纯色填充。
# 3）表头单元格边框是在底部和右侧添加细框，表中和表尾单元格边框是在单元格左侧添加细框。
# 4）同时，表头、表中、表尾的单元格内，对齐方式为水平和垂直方向上的居中对齐。
import os
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Alignment, Side, Border

# 定义表头颜色样式为橙色
header_fill  = PatternFill('solid',fgColor='FF7F24')
# 定义表中颜色样式为淡黄色
content_fill = PatternFill('solid',fgColor='FFFFE0')
# 定义表尾颜色样式为淡桔红色
bottom_fill = PatternFill('solid',fgColor='EE9572')
# 定义对齐样式横向居中、纵向居中
alig = Alignment(horizontal='center',vertical='center')

# 定义边样式为细条
side = Side('thin')
# 定义表头边框样式，有底边和右边
header_border = Border(bottom=side, right=side)
# 定义表中、表尾边框样式，有左边
content_border = Border(left=side)

# 设置文件夹路径
path = '../各部门利润表汇总/'
# 返回当前目录下所有文件名
files = os.listdir(path) 

# 循环文件名列表
for file in files:
    # 拼接文件路径
    file_path= path + file
    # 打开工作簿
    wd = load_workbook(file_path)
    # 打开工作表
    ws = wd.active
    # 调整列宽
    ws.column_dimensions['A'].width=10
    ws.column_dimensions['B'].width=25
    ws.column_dimensions['C'].width=50
    ws.column_dimensions['D'].width=10
    ws.column_dimensions['E'].width=20
    ws.column_dimensions['F'].width=15
    # 循环第一行单元格，调整表头样式
    for cell in ws[1]:
        # 设置单元格填充颜色
        cell.fill=header_fill
        # 设置单元格对齐方式
        cell.alignment=alig
        # 设置单元格边框
        cell.border=header_border


    # 获取最后一行行号
    row_num=ws.max_row


    # 从第二行开始，循环到倒数第二行
    for row in ws.iter_rows(min_row=2,max_row=(row_num-1)):

        # 循环取出单元格，调整表中样式
        for cell in row:
            # 设置单元格填充颜色
            cell.fill=content_fill
            # 设置单元格对齐方式
            cell.alignment=alig
            # 设置单元格边框
            cell.border=content_border
    # 循环最后一行单元格，调整表尾样式
    for cell in ws[row_num]:
        # 设置单元格填充颜色
        cell.fill=bottom_fill
        # 设置单元格对齐方式
        cell.alignment=alig
        # 设置单元格边框
        cell.border=content_border
    # 保存
    wd.save(file_path)
    
