# 导入模块
from sqlite3 import Row
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
# 打开工作表
file_path = './material/事业02部_副本.xlsx'
wb = load_workbook(file_path)
ws = wb.active

# 打印最大行max_row
rows=ws.max_row
print(rows)