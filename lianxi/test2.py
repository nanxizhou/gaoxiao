# 案例 2：计算车间每日产量达标率
# 目标：把【生产计划表.xlsx】工作簿和【工人产量日报表.xlsx】工作簿中对应的数据
# 复制或经过计算记录到【检验记录表模板.xlsx】工作簿中，最后另存为【8月25日检验记录表.xlsx】工作簿
# 请随时查看知识库和案例练习助手，与自己编写代码的步骤和内容比对参考，训练思维
from openpyxl import load_workbook
# 1. “实际产量”的累加
# - 打开【工人产量日报表.xlsx】工作簿，获取【工人产量日报表】工作表
production_wb = load_workbook('../工作/工人产量日报表.xlsx')
production_sheet = production_wb.active
# - 创建空字典，用以存储车间信息
production_dict = {}
# - 遍历【工人产量日报表】工作表中的数据
for row in production_sheet.iter_rows(min_row=3, values_only=True):
    # - 从"工号"中提取出"车间号"
    workshop_num = row[0][:2]
    # - 判断字典是否有车间信息
    if production_dict.get( workshop_num ) == None:
    # - 如果没有。则以“车间号”为键，“产品编号”与“实际产量”组成的字典为值，写入字典
        production_dict[workshop_num] = {row[1]: row[4]}
    # - 如果有。继续判断当前车间中，是否有该产品的信息
    else:
        if production_dict[workshop_num].get(row[1]) !=  None:
    # - 如果有。则累加"实际产量"
            production_dict[workshop_num][row[1]] += row[4]
    # - 如果没有。则以“产品编号”为键，“实际产量”为值，写入字典中
        else:
            production_dict[workshop_num][row[1]] = row[4]
# 2. 数据的写入
# - 打开【检验记录模板.xlsx】工作簿，获取【检验记录模板】工作表
template_wb = load_workbook('../工作/检验记表模板.xlsx')
template_sheet =  template_wb.active
# - 打开【生产计划表.xlsx】工作簿，获取【生产计划表】工作表
plan_wb = load_workbook('../工作/生产计划表.xlsx')
plan_sheet = plan_wb.active
# - 遍历【生产计划表】工作表中的数据
for row in plan_sheet.iter_rows(min_row=3, values_only=True):
    # - 根据“车间号”、“产品编号”，获取产量字典中的“实际产量”
    actual_production = production_dict[row[1]][row[2]]
    # - 计算“目标达成率”，四舍五入并转化成百分比格式
    rate = actual_production/row[3]
    rate = str(round(rate*100, 2)) + '%'
    # - 将【生产计划表】工作表前四列数据，“实际产量”和“目标达成率”添加到【检验记录表模板】工作表中
    template_row = row[:4] + (actual_production, rate)
    template_sheet.append(template_row)
    # - 另存为【8月25日检验记录表.xlsx】工作簿
template_wb.save('../工作/8月25日检验记录表.xlsx')
