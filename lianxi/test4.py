# 案例 4：统计广告投放数据
# 目标：把【销售数据表格.xlsx】工作簿和【账户报表.xlsx】工作簿中对应的数据
# 复制或经过计算记录到【广告投放数据.xlsx】工作簿中
# 最后另存为【2020年9月11日广告投放数据.xlsx】工作簿
# 请随时查看知识库和案例练习助手，与自己编写代码的步骤和内容比对参考，训练思维
from openpyxl import load_workbook

# 1. 【销售数据表格.xlsx】中数据的读取
# - 打开【销售数据表格.xlsx】工作簿，获取【销售数据表格】工作表
sell_wb = load_workbook('../工作/销售数据表格.xlsx')
sell_sheet = sell_wb.active
# - 创建存储销售数据的字典
sell_dict = {'成交单数': 0, '成交额': 0}
# - 遍历【销售数据表格】工作表的数据
for row in sell_sheet.iter_rows(min_row=2, values_only=True):
    # 累计“成交单数”，计算并累加“成交额”
    sell_dict['成交单数'] += 1
# 计算并累加成交额
    sell_dict['成交额'] += row[2] * row[3]
# - 计算“平均客单价”，获取“当天日期”(即“下单日期”)
#     平均客单价
    per_transaction = sell_dict['成交额'] / sell_dict['成交单数']
    # 下单日期
    date = sell_sheet['G2'].value
# 2. 【账户报表.xlsx】中数据的读取与相关指标计算
# - 打开【账户报表.xlsx】工作簿，获取【账户报表】工作表
account_wb = load_workbook('../工作/账户报表.xlsx')
account_sheet = account_wb.active
# - 获取“花费”、“点击量”，计算“PPC”
# 花费
cost = account_sheet['G2'].value
# 点击量
click_amount = account_sheet['E2'].value
# PPC
PPC = cost / click_amount
# - 获取“点击率”、“总加购数”，计算“加购率”并转成百分比格式
# 点击率
click_rate = account_sheet['F2'].value
# 总加购数
add_amount = account_sheet['L2'].value
# 加购率
add_rate = add_amount / click_rate
add_rate = '{}%'.format(round(add_rate * 100, 2))
# - 计算“转化率”并转成百分比格式，计算“成交ROI”
# 转化率
conversions_rate = sell_dict['成交单数'] / click_amount
conversions_rate = '{}%'.format(round(conversions_rate * 100, 2))
# 成交ROI
ROI = sell_dict['成交额'] / cost
ROI = '{}%'.format(round(ROI * 100, 2))
# 3. 【广告投放数据】中数据的写入
# - 根据【广告投放数据】工作表的表头顺序，将上方得到的数据拼接成一行
row_data = [cost, click_amount, PPC, click_rate, add_amount, add_rate, sell_dict['成交单数'], sell_dict['成交额'],
            per_transaction, conversions_rate, ROI]
# - 打开【广告投放数据.xlsx】工作簿，获取【广告投放数据】工作表
advertising_wb = load_workbook('../工作/广告投放数据.xlsx')
advertising_sheet = advertising_wb.active
# - 遍历【广告投放数据】工作表的数据
for row in advertising_sheet.iter_rows(min_row=4):
    # 判断该行中的“日期”是否为“下单日期”
    if row[1].value == date:
        # 如果日期一致。通过循环取出 0-10
        for i in range(11):
            #  将各项数据写到该行中
            row[i+2].value = row_data[i]
# - 另存为【XXX广告投放数据.xlsx】工作簿
advertising_wb.save('../工作/{}广告投放数据.xlsx'.format(date))