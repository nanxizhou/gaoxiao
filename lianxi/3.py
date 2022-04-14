# 案例 10：群发推广邮件
from openpyxl import load_workbook
import smtplib
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart

# 设置邮箱服务器，端口
smtp = smtplib.SMTP_SSL('smtp.qq.com', 465)

# 设置正文内容
content = '''
您好！
附件中为天变公司新产品的介绍，您注意查阅，有什么问题请随时联系！
'''
email_content = MIMEText(content, 'plain', 'utf-8')

# 读取附件【新产品介绍.pdf】文件数据
with open('../工作/新产品介绍.pdf', 'rb') as f:
    file_data = f.read()

# 设置内容类型为附件
attachment = MIMEText(file_data, 'base64', 'utf-8')

# 设置附件标题以及文件类型
attachment.add_header('Content-Disposition',
                      'attachment', filename='新产品介绍.pdf')

# 创建发件邮箱账号列表
mail_list = [
    {'email': 'xxx@qq.com', 'token': 'xxx'},
    {'email': 'xxx@qq.com', 'token': 'xxx'},
]

# 获取【客户名单】工作表
wb = load_workbook('../工作/客户名单.xlsx')
sheet = wb.active

# 创建客户邮箱账号列表
costumer_list = []
# 遍历【客户名单】工作表中数据
for row in sheet.iter_rows(min_row=2, values_only=True):
    # 将数据写入客户邮箱账号列表中
    costumer_list.append(row[3])

# 获取客户邮箱账号数量
costumer_num = sheet.max_row - 1

# 判断客户邮箱账号数量除以6余数是否大于0
if costumer_num % 6 > 0:
    # 如果判断大于0，发件邮箱账号个数等于客户邮箱账号数量整除6的值加1
    account_num = (costumer_num // 6) + 1

else:
    # 否则，发件邮箱账号个数等于客户邮箱账号数量整除6的值
    account_num = costumer_num // 6

# 循环发件邮箱账号个数
for mail_count in range(account_num):
    # 取出发件邮箱账号的邮箱和授权码
    email = mail_list[mail_count]['email']
    token = mail_list[mail_count]['token']

    # 从客户邮箱账号列表中取出单个发件邮箱账号需要发送的客户邮箱账号
    receiver_list = costumer_list[mail_count * 6: (mail_count + 1) * 6]

    # 判断单个发件邮箱账号需要发送的客户邮箱账号数量除以3余数是否大于0
    if len(receiver_list) % 3 > 0:
        # 如果判断大于0，单个发件邮箱账号需要发送的次数等于单个发件邮箱账号需要发送的客户邮箱账号数量整除3的值加1
        send_times = (len(receiver_list) // 3) + 1

    else:
        # 否则，单个发件邮箱账号需要发送的次数等于单个发件邮箱账号需要发送的客户邮箱账号数量整除3的值
        send_times = len(receiver_list) // 3

    # 循环单个发件邮箱账号需要发送的次数
    for i in range(send_times):
        # 取出单次发送的客户邮箱账号
        receivers = receiver_list[i * 3:(i + 1) * 3]

        # 登录发件邮箱账号
        smtp.login(email, token)

        # 设置邮件内容
        msg = MIMEMultipart()

        # 设置发送人、收件人、邮件主题
        msg['From'] = email
        msg['To'] = '尊敬的客户'
        msg['Subject'] = '新产品介绍'

        # 将正文内容添加到邮件内容里
        msg.attach(email_content)

        # 将附件添加到邮件内容里
        msg.attach(attachment)

        # 发送邮件
        smtp.sendmail(email, receivers, msg.as_string())

# 关闭邮箱服务
smtp.quit()
