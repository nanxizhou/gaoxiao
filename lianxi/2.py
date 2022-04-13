from openpyxl import load_workbook

# 获取【工人产量日报表】工作表
production_wb = load_workbook('../工作/工人产量日报表.xlsx')
production_sheet = production_wb.active

# 创建产量字典
production_dict = {}

# 遍历【工人产量日报表】工作表中的数据
for row in production_sheet.iter_rows(min_row=3, values_only=True):
    # 从“工号”中提取出“车间号”（workshop_num)
    workshop_num = row[0][:2]

    # 判断字典中是否有该车间的信息
    if production_dict.get(workshop_num) == None:
        # 以“车间号”为键，“产品编号”与“实际产量 ”组成的字典为值，写入字典中
        production_dict[workshop_num] = {row[1]: row[4]}

    else:
        # 判断当前车间中，是否有该产品的信息
        if production_dict[workshop_num].get(row[1]) != None:
            # 累加“实际产量”
            production_dict[workshop_num][row[1]] += row[4]

        # 如不存在该产品信息，则以“产品编号”为键，“实际产量”为值，写入字典中
        else:
            production_dict[workshop_num][row[1]] = row[4]

# 分别获取【检验记录表模板】、【生产计划表】工作表
template_wb = load_workbook('../工作/检验记录表模板.xlsx')
template_sheet = template_wb.active

plan_wb = load_workbook('../工作/生产计划表.xlsx')
plan_sheet = plan_wb.active

# 遍历【生产计划表】工作表中的数据
for row in plan_sheet.iter_rows(min_row=3, values_only=True):
    # 根据“车间号”、“产品编号”，获取产量字典中的“实际产量”
    actual_production = production_dict[row[1]][row[2]]

    # 计算“目标达成率”
    rate = actual_production / row[3]

    # 将“目标达成率”保留两位小数，并转成百分比格式
    rate = str(round(rate*100, 2)) + '%'

    # 根据【检验记录表模板】工作表的表头，将数据写入【检验记录表模板】工作表中
    template_row = row[:4] + (actual_production, rate)
    template_sheet.append(template_row)

# 将【检验记录表模板.xlsx】工作簿另存为【8月25日检验记录表.xlsx】工作簿
template_wb.save('../工作/8月25日检验记录表.xlsx')