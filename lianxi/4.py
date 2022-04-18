# 案例 4：统计广告投放数据
from openpyxl import load_workbook

# 获取【销售数据表格】工作表
sell_wb = load_workbook('../工作/销售数据表格.xlsx')
sell_sheet = sell_wb.active

# 创建销售数据字典
sell_dict = {'成交单数': 0, '成交额': 0}
# 遍历【销售数据表格】表中的数据
for row in sell_sheet.iter_rows(min_row=2, values_only=True):
    # 累计成交单数
    sell_dict['成交单数'] += 1
    # 计算并累加成交额
    sell_dict['成交额'] += row[2] * row[3]

# 计算“平均客单价”
per_transaction = sell_dict['成交额'] / sell_dict['成交单数']

# 获取当天日期，即【销售数据表格】中的“下单日期”
date = sell_sheet['G2'].value

# 获取【账户报表】工作表
account_wb = load_workbook('../工作/账户报表.xlsx')
account_sheet = account_wb.active

# 获取“花费”、“点击量”，并计算“PPC”
cost = account_sheet['G2'].value
click_amount = account_sheet['E2'].value
PPC = cost / click_amount

# 获取“点击率”、“总加购数”，计算“加购率”并转成百分比格式
click_rate = account_sheet['F2'].value
add_amount = account_sheet['L2'].value
add_rate = add_amount / click_amount
add_rate = '{}%'.format(round(add_rate * 100, 2))

# 计算“转化率”,并转成百分比格式
conversions_rate = sell_dict['成交单数'] / click_amount
conversions_rate = '{}%'.format(round(conversions_rate * 100, 2))

# 计算“成交ROI”
ROI = sell_dict['成交额'] / cost

# 根据【广告投放数据】表头，将上面数据拼接成一行
row_data = [cost, click_amount, PPC, click_rate, add_amount, add_rate, sell_dict['成交单数'], sell_dict['成交额'],
            per_transaction, conversions_rate, ROI]

# 获取【广告投放数据】工作表
advertising_wb = load_workbook('../工作/广告投放数据.xlsx')
advertising_sheet = advertising_wb.active

# 遍历【广告投放数据】表中数据
for row in advertising_sheet.iter_rows(min_row=3):
    # 判断当前“日期”是否为“下单日期”
    if row[1].value == date:

        # 通过循环取出0-10
        for i in range(11):
            # 修改对应单元格的值
            row[i+2].value = row_data[i]

# 将【广告投放数据.xlsx】另存为【“当天日期”+广告投放数据.xlsx】
advertising_wb.save('../工作/{}广告投放数据.xlsx'.format(date))